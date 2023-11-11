from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

# code to style the excel sheets 


def style_worksheet(filename):
	print('styling')
	wb = load_workbook(filename)
	ws = wb.active # first sheet

	start_row = 3
	end_row = 4
	start_col = 1
	
	center_cell_text(ws, start_row, start_col)
	adjust_column_widths(ws, start_row, end_row, start_col)
	create_table_borders(ws,start_row, end_row, start_col)
	style_table(ws, start_row)

	wb.save(filename)


def style_table(worksheet, start_row):
	style = TableStyleInfo(name="TableStyleMedium2",
						   showFirstColumn=False,
						   showLastColumn=False,
					       showRowStripes=True,
					       showColumnStripes=False)

	data_table = Table(displayName="table1", ref=f'$A{start_row}:{get_column_letter(worksheet.max_column)}{worksheet.max_row}')


	data_table.tableStyleInfo = style

	worksheet.add_table(data_table)

	# center column titles and values in table


def center_cell_text(worksheet, start_row, start_col):
	for row in range(start_row, worksheet.max_row + 1):
		for col in range(start_col, worksheet.max_column + 1):
			worksheet.cell(row=row, column=col).alignment = Alignment(horizontal='center')


def adjust_column_widths(worksheet, start_row, end_row, start_col):
	column_widths = []
	PADDING = 5

	for row in range(start_row, end_row + 1):
		for col in range(start_col, worksheet.max_column + 1):
			char_len = len(worksheet.cell(row=row, column=col).value) + PADDING
			if len(column_widths) > col:
				
				if char_len > column_widths[col]:
					column_widths[col] = char_len
			else:
				column_widths.append(char_len)

	for col, column_width in enumerate(column_widths, 1):
		worksheet.column_dimensions[get_column_letter(col)].width = column_width

			

def create_table_borders(worksheet, start_row, end_row_header, start_col):
	
	# add thick border to top of table header
	# add_border_to_cells(worksheet, start_row, start_row, start_col, worksheet.max_column, "thick", "top")

	# add thick border to bottom of table header
	add_border_to_cells(worksheet, end_row_header, end_row_header, start_col, worksheet.max_column, "thick", "bottom")

	# add thick border to left side of table
	# add_border_to_cells(worksheet, start_row, worksheet.max_row, start_col, start_col, "thick", "left")

	# add thick border to right side
	# add_border_to_cells(worksheet, start_row, worksheet.max_row, worksheet.max_column, worksheet.max_column, "thick", "right")

	



def add_border_to_cells(worksheet, start_row, end_row, start_col, end_col, style, location):

	# add border to cells 
	for row in range(start_row, end_row + 1):
		for col in range(start_col, end_col + 1):
			
			cell = worksheet.cell(row=row, column=col)
			
			add_cell_border(cell, Side(border_style=style), location)



def add_cell_border(cell, border_style, location):
	if location == "top":
		cell.border = Border(top=border_style)
	elif location == "left":
		cell.border = Border(left=border_style)
	elif location == "right":
		cell.border = Border(right=border_style)
	elif location == "bottom":
		cell.border = Border(bottom=border_style)
	
	



